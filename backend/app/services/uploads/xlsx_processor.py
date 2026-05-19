"""
XLSX template generator and parsers.

Tab A — Entitlement Update (all fields in one sheet):
  LOCKED (grey):    ENT_ID | SW_ID | Canonical Name | Metric | Current Status | PO Number
  EDITABLE (white): Contract Name | License Type | Entitled Count | In-Use Count
                    | Unit Cost (INR) | Annual Cost (INR) | Notes

Tab B — License Discovery (device-level usage data):
  LOCKED (grey):    ENT_ID | SW_ID | Contract Software Name
  EDITABLE (white): Application Tagged | Data Source | Device Type (endpoint/server)
                    | Device ID | OS | Version | Last Seen (YYYY-MM-DD) | Region
"""
import io
import hashlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


_HEADER_FILL = PatternFill("solid", fgColor="1A2E5A")   # DRL navy
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_LOCKED_FILL = PatternFill("solid", fgColor="F0F2F5")   # light grey = read-only hint


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def generate_template(entitlements: list[dict]) -> bytes:
    """
    Build an XLSX workbook with two sheets pre-populated from entitlements list.
    Each dict must have: ent_id, sw_id, canonical_name, metric_name, status,
                         contract_name, license_type, entitled_count,
                         unit_cost_inr, annual_cost_inr, po_number,
                         in_use_count, notes
    Returns raw bytes of the .xlsx file.
    """
    wb = Workbook()

    # ── Tab A ──────────────────────────────────────────────────────────────────
    # Cols 1-5 locked (reference), cols 6-12 editable
    ws_a = wb.active
    ws_a.title = "Tab A - Metadata"
    headers_a = [
        "ENT_ID", "SW_ID", "Canonical Name", "Metric", "Current Status", "PO Number",  # locked (1-6)
        "Contract Name", "License Type (subscription/perpetual)",                        # editable (7-8)
        "Entitled Count", "In-Use Count",                                                # editable (9-10)
        "Unit Cost (INR)", "Annual Cost (INR)", "Notes",                                 # editable (11-13)
    ]
    ws_a.append(headers_a)
    _style_header(ws_a, 1, len(headers_a))

    for ent in entitlements:
        ws_a.append([
            ent.get("ent_id", ""),
            ent.get("sw_id", ""),
            ent.get("canonical_name", ""),
            ent.get("metric_name", ""),
            ent.get("status", ""),
            ent.get("po_number") or "",
            ent.get("contract_name", ""),
            ent.get("license_type", ""),
            ent.get("entitled_count") or "",
            ent.get("in_use_count") or "",
            ent.get("unit_cost_inr") or "",
            ent.get("annual_cost_inr") or "",
            ent.get("notes") or "",
        ])

    # Grey out locked reference columns (1-6)
    for row in ws_a.iter_rows(min_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.fill = _LOCKED_FILL

    # ── Tab B — License Discovery ─────────────────────────────────────────────
    # Locked (cols 1-3): ENT_ID | SW_ID | Contract Software Name
    # Editable (cols 4-12): Application Tagged | Data Source | Device Type
    #   | Device ID | OS | Version | Last Seen (YYYY-MM-DD) | Site | Region
    ws_b = wb.create_sheet("Tab B - License Discovery")
    headers_b = [
        "ENT_ID", "SW_ID", "Contract Software Name",              # locked (1-3)
        "Application Tagged", "Data Source",                       # editable (4-5)
        "Device Type (endpoint/server)", "Device ID",              # editable (6-7)
        "OS", "Version", "Last Seen (YYYY-MM-DD)", "Site", "Region",  # editable (8-12)
    ]
    ws_b.append(headers_b)
    _style_header(ws_b, 1, len(headers_b))

    for ent in entitlements:
        ws_b.append([
            ent.get("ent_id", ""),
            ent.get("sw_id", ""),
            ent.get("contract_name", "") or ent.get("canonical_name", ""),
            "",  # Application Tagged — user fills
            "",  # Data Source — user fills
            "",  # Device Type — user fills
            "",  # Device ID — user fills
            "",  # OS — user fills
            "",  # Version — user fills
            "",  # Last Seen — user fills
            "",  # Site — user fills
            "",  # Region — user fills
        ])

    # Grey out locked reference columns (1-3)
    for row in ws_b.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.fill = _LOCKED_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_tab_a(data: bytes) -> list[dict]:
    """
    Parse Tab A sheet. Returns list of dicts with updatable keys.
    Column layout (0-indexed):
      0  ENT_ID (key, locked)
      1  SW_ID (locked)
      2  Canonical Name (locked)
      3  Metric (locked)
      4  Current Status (locked)
      5  PO Number (locked)
      6  Contract Name (editable)
      7  License Type (editable — normalised to lowercase)
      8  Entitled Count (editable)
      9  In-Use Count (editable)
      10 Unit Cost INR (editable)
      11 Annual Cost INR (editable)
      12 Notes (editable)
    Skips rows where ENT_ID is blank.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Tab A - Metadata"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = []
    for row in rows:
        if not row or not row[0]:
            continue
        license_type = str(row[7]).strip().lower() if row[7] else None
        if license_type not in ("subscription", "perpetual"):
            license_type = None  # ignore invalid values
        entitled   = int(row[8])  if row[8]  is not None else None
        in_use     = int(row[9])  if row[9]  is not None else None
        unit_cost  = int(row[10]) if row[10] is not None else None
        annual_raw = int(row[11]) if row[11] is not None else None
        # Auto-calc annual_cost when blank but entitled × unit_cost available
        annual_cost = annual_raw if annual_raw is not None else (
            entitled * unit_cost if (entitled is not None and unit_cost is not None) else None
        )
        result.append({
            "ent_id":          str(row[0]).strip(),
            "sw_id":           str(row[1]).strip() if row[1] else None,
            "contract_name":   str(row[6]).strip() if row[6] else None,
            "license_type":    license_type,
            "entitled_count":  entitled,
            "in_use_count":    in_use,
            "unit_cost_inr":   unit_cost,
            "annual_cost_inr": annual_cost,
            "notes":           str(row[12]).strip() if row[12] else None,
        })
    return result


def parse_tab_b_discovery(data: bytes) -> list[dict]:
    """
    Parse Tab B (License Discovery) sheet.
    Column layout (0-indexed):
      0  ENT_ID (locked — reference/lookup)
      1  SW_ID (locked — reference/lookup)
      2  Contract Software Name (locked — becomes discovery.contract_name)
      3  Application Tagged (editable)
      4  Data Source (editable)
      5  Device Type (editable — normalised to endpoint/server)
      6  Device ID (editable)
      7  OS (editable)
      8  Version (editable)
      9  Last Seen YYYY-MM-DD (editable)
      10 Site (editable)
      11 Region (editable)
    Skips rows where both ENT_ID and Contract Software Name are blank.
    """
    from datetime import datetime as _dt
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Tab B - License Discovery"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = []
    for row in rows:
        if not row:
            continue
        ent_id        = str(row[0]).strip() if row[0] else None
        sw_id         = str(row[1]).strip() if row[1] else None
        contract_name = str(row[2]).strip() if row[2] else None
        if not ent_id and not contract_name:
            continue  # blank row

        raw_device_type = str(row[5]).strip().lower() if row[5] else "endpoint"
        device_type = raw_device_type if raw_device_type in ("endpoint", "server") else "endpoint"

        raw_date = row[9]
        last_seen = None
        if raw_date:
            try:
                if hasattr(raw_date, "date"):
                    last_seen = raw_date.date()
                else:
                    last_seen = _dt.strptime(str(raw_date).strip(), "%Y-%m-%d").date()
            except ValueError:
                pass

        result.append({
            "ent_id":            ent_id,
            "sw_id":             sw_id,
            "contract_name":     contract_name,
            "application_tagged": str(row[3]).strip() if row[3] else None,
            "data_source":        str(row[4]).strip() if row[4] else None,
            "device_type":        device_type,
            "device_id":          str(row[6]).strip() if row[6] else None,
            "os":                 str(row[7]).strip() if row[7] else None,
            "version":            str(row[8]).strip() if row[8] else None,
            "last_seen":          last_seen,
            "site":               str(row[10]).strip() if row[10] else None,
            "region":             str(row[11]).strip() if row[11] else None,
        })
    return result


def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def xls_to_xlsx(data: bytes) -> bytes:
    """
    Convert legacy .xls bytes to .xlsx bytes using xlrd + openpyxl.
    Preserves sheet names and cell values; formatting is not preserved
    (not needed — we only care about data values for parsing).
    """
    import xlrd
    wb_src = xlrd.open_workbook(file_contents=data)
    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)
    for sheet_name in wb_src.sheet_names():
        ws_src = wb_src.sheet_by_name(sheet_name)
        ws_dst = wb_dst.create_sheet(sheet_name)
        for row_idx in range(ws_src.nrows):
            ws_dst.append([ws_src.cell_value(row_idx, col_idx) for col_idx in range(ws_src.ncols)])
    buf = io.BytesIO()
    wb_dst.save(buf)
    return buf.getvalue()
