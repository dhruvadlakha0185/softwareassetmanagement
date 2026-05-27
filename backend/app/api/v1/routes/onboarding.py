import io
from uuid import UUID
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.core.database import get_db
from app.api.deps import get_current_user, require_role
from app.models.catalog import SoftwareCatalog, SoftwareAlias
from app.models.contracts import Contract, Entitlement, OnboardingDraft, EntitlementPriceSchedule, EntitlementDoaContact
from app.schemas.onboarding import (
    DraftSave, DraftOut, PublishPayload, PublishOut,
    MultiPublishPayload, MultiPublishOut, MultiPublishCreated,
)
from app.services.ai.contract_extractor import extract_contract_text, call_openai
from app.services.ai.notes_generator import generate_entitlement_notes

router = APIRouter(prefix="/onboarding", tags=["onboarding"])
auth = Depends(get_current_user)


def _publisher_prefix(publisher: str | None) -> str:
    """Derive a 2-4 char uppercase prefix from a publisher name.

    Rules (applied in order):
    - Known overrides first (SAP SE → SAP, etc.)
    - Split on spaces/punctuation, take first letter of each word up to 3 words
    - Collapse to uppercase, max 4 chars, min 2 chars
    """
    OVERRIDES = {
        "microsoft": "MS",
        "adobe": "ADO",
        "sap se": "SAP",
        "sap": "SAP",
        "oracle": "ORC",
        "salesforce": "SFD",
        "servicenow": "SNW",
        "veeva systems": "VVA",
        "veeva": "VVA",
        "ibm": "IBM",
        "atlassian": "ATL",
        "autodesk": "ADS",
        "ansys": "ANS",
        "siemens": "SIE",
        "aveva": "AVA",
        "hexagon": "HEX",
        "opentext": "OTX",
        "broadcom": "BDC",
        "citrix": "CTX",
        "vmware": "VMW",
        "tableau": "TBL",
        "mimecast": "MMC",
        "crowdstrike": "CRS",
        "zscaler": "ZSC",
        "palo alto networks": "PAN",
        "palo alto": "PAN",
        "fortinet": "FTN",
        "trend micro": "TMI",
        "zoom": "ZOM",
        "slack": "SLK",
        "box": "BOX",
        "dropbox": "DBX",
        "github": "GHB",
        "gitlab": "GLB",
        "jira": "JRA",
        "docusign": "DCU",
        "workday": "WKD",
        "successfactors": "SCF",
        "sap successfactors": "SCF",
    }
    if not publisher:
        return "SW"
    key = publisher.strip().lower()
    if key in OVERRIDES:
        return OVERRIDES[key]
    # Build from initials of each word (max 3 words)
    import re
    words = re.split(r"[\s\-,./]+", key)
    words = [w for w in words if w and not w.isdigit()]
    if not words:
        return "SW"
    prefix = "".join(w[0] for w in words[:3]).upper()
    return prefix if len(prefix) >= 2 else (prefix + words[0][1:3].upper())[:4]


async def _next_sw_id(db: AsyncSession, publisher: str | None = None) -> str:
    prefix = _publisher_prefix(publisher)
    pattern = f"{prefix}-%"
    result = await db.execute(
        select(func.max(SoftwareCatalog.sw_id)).where(SoftwareCatalog.sw_id.like(pattern))
    )
    max_id = result.scalar_one_or_none()
    if max_id:
        try:
            n = int(max_id.split("-")[1]) + 1
        except (IndexError, ValueError):
            n = 1
    else:
        n = 1
    while True:
        candidate = f"{prefix}-{n:03d}"
        if not await db.get(SoftwareCatalog, candidate):
            return candidate
        n += 1


async def _next_ent_id(db: AsyncSession) -> str:
    result = await db.execute(
        select(func.max(Entitlement.ent_id)).where(Entitlement.ent_id.like("ENT-%"))
    )
    max_id = result.scalar_one_or_none()
    if max_id:
        try:
            n = int(max_id.split("-")[1]) + 1
        except (IndexError, ValueError):
            n = 1
    else:
        n = 1
    return f"ENT-{n:03d}"


# ── AI Extraction ──────────────────────────────────────────────────────────────

@router.post("/extract")
async def extract_contract(
    file: UploadFile = File(...),
    current_user=auth,
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    data = await file.read()
    try:
        text = extract_contract_text(file.filename, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    result = await call_openai(text)
    return result


# ── Drafts ─────────────────────────────────────────────────────────────────────

@router.get("/drafts", response_model=list[DraftOut])
async def list_drafts(current_user=auth, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(OnboardingDraft).where(OnboardingDraft.user_id == current_user.id)
    )
    return [DraftOut.model_validate(d) for d in result.scalars().all()]


@router.post("/drafts", response_model=DraftOut, status_code=201)
async def create_draft(
    body: DraftSave,
    current_user=auth,
    db: AsyncSession = Depends(get_db),
):
    draft = OnboardingDraft(user_id=current_user.id, **body.model_dump())
    db.add(draft)
    await db.commit()
    await db.refresh(draft)
    return DraftOut.model_validate(draft)


@router.get("/drafts/{draft_id}", response_model=DraftOut)
async def get_draft(
    draft_id: UUID,
    current_user=auth,
    db: AsyncSession = Depends(get_db),
):
    draft = await db.get(OnboardingDraft, draft_id)
    if not draft or draft.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Draft not found")
    return DraftOut.model_validate(draft)


@router.put("/drafts/{draft_id}", response_model=DraftOut)
async def update_draft(
    draft_id: UUID,
    body: DraftSave,
    current_user=auth,
    db: AsyncSession = Depends(get_db),
):
    draft = await db.get(OnboardingDraft, draft_id)
    if not draft or draft.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Draft not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(draft, k, v)
    await db.commit()
    await db.refresh(draft)
    return DraftOut.model_validate(draft)


@router.delete("/drafts/{draft_id}", status_code=204)
async def delete_draft(
    draft_id: UUID,
    current_user=auth,
    db: AsyncSession = Depends(get_db),
):
    draft = await db.get(OnboardingDraft, draft_id)
    if not draft or draft.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Draft not found")
    await db.delete(draft)
    await db.commit()


# ── Publish ────────────────────────────────────────────────────────────────────

@router.post("/publish", response_model=PublishOut, status_code=201)
async def publish_onboarding(
    body: PublishPayload,
    current_user=auth,
    db: AsyncSession = Depends(get_db),
):
    # Resolve or create SW catalog entry
    if body.sw_id:
        sw = await db.get(SoftwareCatalog, body.sw_id)
        if not sw:
            raise HTTPException(status_code=404, detail=f"SW entry {body.sw_id} not found")
        sw_id = body.sw_id
    else:
        existing = (await db.execute(
            select(SoftwareCatalog).where(SoftwareCatalog.primary_sw_name == body.primary_sw_name)
        )).scalar_one_or_none()
        if existing:
            sw_id = existing.sw_id
        else:
            sw_id = await _next_sw_id(db, body.publisher)
            sw = SoftwareCatalog(
                sw_id=sw_id,
                primary_sw_name=body.primary_sw_name,
                publisher=body.publisher,
                category_id=body.category_id,
                sub_category_id=body.sub_category_id,
                gxp_flag=body.gxp_flag,
                vendor_id=body.vendor_id,
                vendor_risk=body.vendor_risk,
                deployment=body.deployment,
                region_id=body.region_id,
                app_owner_id=body.app_owner_id,
                notes=body.notes,
                onboarded_date=date.today(),
                created_by=current_user.id,
            )
            db.add(sw)
            await db.flush()

    # Add aliases (Step 6)
    for alias_name in body.aliases:
        existing_alias = (await db.execute(
            select(SoftwareAlias).where(
                SoftwareAlias.sw_id == sw_id,
                SoftwareAlias.alias_name == alias_name,
            )
        )).scalar_one_or_none()
        if not existing_alias:
            from datetime import datetime as _dt
            db.add(SoftwareAlias(sw_id=sw_id, alias_name=alias_name, date_mapped=_dt.utcnow()))

    # Create contract record
    contract = Contract(
        sw_id=sw_id,
        po_number=body.po_number,
        clm_id=body.clm_id,
        vendor_id=body.vendor_id,
        reseller=body.reseller,
        start_date=body.start_date,
        end_date=body.end_date,
        total_value_inr=body.total_value_inr,
        auto_renewal_clause=body.auto_renewal_clause,
        file_name=body.file_name,
        file_path=body.file_path,
        storage_backend="supabase",
        created_by=current_user.id,
    )
    db.add(contract)
    await db.flush()

    # Create entitlement for each line item
    ent_ids = []
    for item in body.line_items:
        ent_id = await _next_ent_id(db)
        ent = Entitlement(
            ent_id=ent_id,
            sw_id=sw_id,
            contract_id=contract.id,
            contract_name=item.contract_name,
            license_type_id=item.license_type_id,
            entitled_count=item.entitled_count,
            unit_cost=item.unit_cost,
            annual_cost=item.annual_cost,
            notes=item.notes,
            region_id=item.region_id,
            discovery_source_id=item.discovery_source_id,
            usage_method_id=item.usage_method_id,
            app_owner_id=item.app_owner_id,
            status="ACTIVE",
        )
        db.add(ent)
        ent_ids.append(ent_id)

    # Audit
    from app.services.audit_logger import log_event
    await log_event(
        db, current_user.id, "SOFTWARE_ONBOARDED", "software_catalog", sw_id,
        sw_id=sw_id,
        after={"primary_sw_name": body.primary_sw_name, "contract_id": str(contract.id), "ent_ids": ent_ids},
        is_gxp=(body.gxp_flag != "no"),
    )

    await db.commit()
    return PublishOut(sw_id=sw_id, contract_id=contract.id, ent_ids=ent_ids)


@router.post("/multi-publish", response_model=MultiPublishOut, status_code=201)
async def multi_publish(
    body: MultiPublishPayload,
    current_user=Depends(require_role(["COE_ADMIN"])),
    db: AsyncSession = Depends(get_db),
):
    """
    New-wizard publish: one contract → multiple SW+ENT pairs.
    Each line item gets its own Contract record (sharing PO/CLM/dates),
    its own SW_ID (new or existing), and its own ENT_ID.
    """
    if not body.line_items:
        raise HTTPException(status_code=400, detail="No line items provided")

    # ── Batch-resolve lookup names for AI notes context ───────────────────
    from app.models.masters import Category, SubCategory, LicenseMetric, LicenseType, Region

    _cat_ids = {i.category_id for i in body.line_items if i.category_id}
    _sub_ids = {i.sub_category_id for i in body.line_items if i.sub_category_id}
    _metric_ids = {i.metric_id for i in body.line_items if i.metric_id}
    _lt_ids = {i.license_type_id for i in body.line_items if i.license_type_id}

    cat_names: dict = {}
    sub_names: dict = {}
    metric_names: dict = {}
    lt_names: dict = {}

    if _cat_ids:
        rows = (await db.execute(select(Category).where(Category.id.in_(_cat_ids)))).scalars().all()
        cat_names = {r.id: r.name for r in rows}
    if _sub_ids:
        rows = (await db.execute(select(SubCategory).where(SubCategory.id.in_(_sub_ids)))).scalars().all()
        sub_names = {r.id: r.name for r in rows}
    if _metric_ids:
        rows = (await db.execute(select(LicenseMetric).where(LicenseMetric.id.in_(_metric_ids)))).scalars().all()
        metric_names = {r.id: r.name for r in rows}
    if _lt_ids:
        rows = (await db.execute(select(LicenseType).where(LicenseType.id.in_(_lt_ids)))).scalars().all()
        lt_names = {r.id: r.license_type for r in rows}

    # Region name → UUID map (used to set region_id FK from free-text region list)
    all_regions = (await db.execute(select(Region))).scalars().all()
    region_name_to_id: dict = {r.name: r.id for r in all_regions}

    created: list[MultiPublishCreated] = []
    skipped: list[str] = []

    for item in body.line_items:
        try:
            # ── Resolve SW entry ──────────────────────────────────────────
            is_new_sw = False

            # ── Generate notes if blank ───────────────────────────────────
            notes = item.notes or None
            if not notes:
                _ctx = {
                    "primary_sw_name": item.primary_sw_name,
                    "publisher": item.publisher or body.vendor_name,
                    "contract_name": item.contract_name,
                    "category_name": cat_names.get(item.category_id) if item.category_id else None,
                    "sub_category_name": sub_names.get(item.sub_category_id) if item.sub_category_id else None,
                    "license_type_name": lt_names.get(item.license_type_id) if item.license_type_id else None,
                    "deployment": item.deployment,
                    "gxp_flag": item.gxp_flag,
                    "entitled_count": item.entitled_count,
                    "metric_name": metric_names.get(item.metric_id) if item.metric_id else None,
                    "business_units": item.business_units,
                    "regions": item.regions,
                    "vendor_name": body.vendor_name,
                    "start_date": str(body.start_date) if body.start_date else None,
                    "end_date": str(body.end_date) if body.end_date else None,
                    "annual_cost": item.annual_cost,
                    "currency": body.currency or "INR",
                    "auto_renewal_clause": body.auto_renewal_clause,
                }
                notes = await generate_entitlement_notes(_ctx)

            if item.sw_id:
                sw = await db.get(SoftwareCatalog, item.sw_id)
                if not sw:
                    skipped.append(f"'{item.contract_name}': SW_ID {item.sw_id} not found")
                    continue
                sw_id = sw.sw_id
            else:
                existing = (await db.execute(
                    select(SoftwareCatalog).where(
                        SoftwareCatalog.primary_sw_name == item.primary_sw_name
                    )
                )).scalar_one_or_none()
                if existing:
                    sw_id = existing.sw_id
                else:
                    publisher = item.publisher or body.vendor_name
                    sw_id = await _next_sw_id(db, publisher)
                    first_region_id = None
                    if item.regions:
                        first_region_id = region_name_to_id.get(item.regions[0])
                    sw_entry = SoftwareCatalog(
                        sw_id=sw_id,
                        primary_sw_name=item.primary_sw_name,
                        publisher=publisher,
                        category_id=item.category_id,
                        sub_category_id=item.sub_category_id,
                        gxp_flag=item.gxp_flag,
                        vendor_risk=item.vendor_risk,
                        deployment=item.deployment,
                        region_id=first_region_id,
                        app_owner_id=item.app_owner_id,
                        secondary_owner_id=item.secondary_owner_id,
                        notes=notes,
                        onboarded_date=date.today(),
                    )
                    db.add(sw_entry)
                    if item.aliases:
                        from datetime import datetime as _dt
                        for alias in item.aliases:
                            if alias.strip():
                                db.add(SoftwareAlias(sw_id=sw_id, alias_name=alias.strip(), date_mapped=_dt.utcnow()))
                    is_new_sw = True

            # ── Create Contract (one per line item, shared header fields) ─
            contract = Contract(
                sw_id=sw_id,
                po_number=body.po_number,
                clm_id=body.clm_id,
                reseller=body.reseller,
                start_date=body.start_date,
                end_date=body.end_date,
                total_value_inr=body.total_value_inr,
                auto_renewal_clause=body.auto_renewal_clause,
                renewal_alert_extra_days=body.renewal_alert_extra_days or None,
                business_units=item.business_units or None,
                currency=body.currency or "INR",
                storage_backend="local",
                created_by=current_user.id,
            )
            db.add(contract)
            await db.flush()

            # ── Create Entitlement ────────────────────────────────────────
            max_ent = (await db.execute(
                select(func.max(Entitlement.ent_id)).where(Entitlement.ent_id.like("ENT-%"))
            )).scalar_one_or_none()
            if max_ent:
                try:
                    ent_n = int(max_ent.split("-")[1]) + 1
                except (IndexError, ValueError):
                    ent_n = 1
            else:
                ent_n = 1
            while True:
                ent_candidate = f"ENT-{ent_n:03d}"
                if not await db.get(Entitlement, ent_candidate):
                    ent_id = ent_candidate
                    break
                ent_n += 1

            ent_region_id = None
            if item.regions:
                ent_region_id = region_name_to_id.get(item.regions[0])
            ent = Entitlement(
                ent_id=ent_id,
                sw_id=sw_id,
                contract_id=contract.id,
                contract_name=item.contract_name,
                metric_id=item.metric_id,
                license_type_id=item.license_type_id,
                entitled_count=item.entitled_count,
                in_use_count=0,
                unit_cost=item.unit_cost,
                annual_cost=item.annual_cost,
                notes=notes,
                vendor_id=contract.vendor_id,
                region_id=ent_region_id,
                regions_json=item.regions or None,
                business_units=item.business_units or None,
                discovery_source_id=item.discovery_source_id,
                usage_method_id=item.usage_method_id,
                app_owner_id=item.app_owner_id,
                secondary_owner_id=item.secondary_owner_id,
                status="ACTIVE",
            )
            db.add(ent)
            await db.flush()

            # ── Insert per-entitlement DOA contacts ───────────────────────
            for doa_id in set(item.doa_contact_ids):
                db.add(EntitlementDoaContact(ent_id=ent_id, doa_contact_id=doa_id))

            # ── Insert price schedule rows (multi-year contracts) ─────────
            if item.price_schedule:
                yr1 = next((s for s in item.price_schedule if s.year_number == 1), None)
                if yr1:
                    ent.unit_cost = yr1.unit_cost
                    ent.annual_cost = yr1.annual_cost
                    ent.entitled_count = yr1.entitled_count
                for sched in item.price_schedule:
                    db.add(EntitlementPriceSchedule(
                        ent_id=ent_id,
                        year_number=sched.year_number,
                        effective_from=sched.effective_from,
                        effective_to=sched.effective_to,
                        entitled_count=sched.entitled_count,
                        unit_cost=sched.unit_cost,
                        annual_cost=sched.annual_cost,
                    ))

            # Resolve final cost values (Year 1 schedule may have overridden item values)
            final_count = ent.entitled_count
            final_unit = ent.unit_cost
            final_annual = ent.annual_cost
            created.append(MultiPublishCreated(
                sw_id=sw_id,
                ent_id=ent_id,
                contract_id=contract.id,
                primary_sw_name=item.primary_sw_name,
                contract_name=item.contract_name,
                is_new_sw=is_new_sw,
                license_type_name=lt_names.get(item.license_type_id) if item.license_type_id else None,
                metric_name=metric_names.get(item.metric_id) if item.metric_id else None,
                entitled_count=final_count,
                unit_cost=final_unit,
                annual_cost=final_annual,
            ))

        except Exception as e:
            skipped.append(f"'{item.contract_name}': {e}")

    from app.services.audit_logger import log_event
    await log_event(
        db, current_user.id, "SOFTWARE_ONBOARDED", "software_catalog",
        str(created[0].sw_id) if created else "none",
        after={"items_created": len(created), "po_number": body.po_number},
        is_gxp=any(i.gxp_flag != "no" for i in body.line_items),
    )
    await db.commit()
    return MultiPublishOut(created=created, skipped=skipped)


# ── Bulk Onboarding ───────────────────────────────────────────────────────────

def _generate_bulk_template(db_lists: dict) -> bytes:
    """
    Build a two-tab XLSX bulk onboarding template.
    db_lists keys: license_types, metrics, regions, categories,
                   sub_categories, discovery_sources, usage_methods
                   (each a list[str] fetched from DB at request time)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    _MANDATORY_FILL = PatternFill("solid", fgColor="C0392B")  # dark red
    _OPTIONAL_FILL  = PatternFill("solid", fgColor="1A2E5A")  # DRL navy
    _SECTION_FILL   = PatternFill("solid", fgColor="2C3E50")  # dark slate
    _WHITE_BOLD     = Font(color="FFFFFF", bold=True, size=10)
    _SECTION_FONT   = Font(color="FFFFFF", bold=True, size=9, italic=True)
    _EXAMPLE_FONT   = Font(italic=True, color="888888", size=9)
    _CENTER         = Alignment(horizontal="center")

    BIZ_UNITS = [
        "All departments", "Finance", "Commercial", "Management", "SCM",
        "Manufacturing", "QA/QC", "Regulatory", "R&D", "QC Labs",
        "Analytical Dev", "Engineering", "Regulatory Affairs", "Drug Safety",
        "Marketing", "Medical", "IT", "IT Security", "SOC",
        "Procurement", "HR", "Training",
    ]

    wb = Workbook()

    # ── _Lists hidden sheet ───────────────────────────────────────────────────
    ws_lists = wb.active
    ws_lists.title = "_Lists"
    lists_data = [
        ("LicenseTypes",    db_lists.get("license_types", [])),
        ("Metrics",         db_lists.get("metrics", [])),
        ("BusinessUnits",   BIZ_UNITS),
        ("Regions",         db_lists.get("regions", [])),
        ("Categories",      db_lists.get("categories", [])),
        ("SubCategories",   db_lists.get("sub_categories", [])),
        ("DiscoverySources",db_lists.get("discovery_sources", [])),
        ("UsageMethods",    db_lists.get("usage_methods", [])),
    ]
    for col_idx, (header, values) in enumerate(lists_data, 1):
        ws_lists.cell(row=1, column=col_idx, value=header)
        for row_idx, val in enumerate(values, 2):
            ws_lists.cell(row=row_idx, column=col_idx, value=val)
    ws_lists.sheet_state = "hidden"

    # ── Tab 1: Contract Information ───────────────────────────────────────────
    ws1 = wb.create_sheet("Contract Information")

    HEADERS_1 = [
        ("Vendor / Publisher Name *", True),
        ("PO Number *",               True),
        ("Contract Name *",           True),
        ("Contract Start Date *",     True),
        ("Contract End Date *",       True),
        ("CLM ID",                    False),
        ("Auto-Renewal Clause",       False),
        ("Currency",                  False),
    ]
    for col_idx, (header, mandatory) in enumerate(HEADERS_1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = _MANDATORY_FILL if mandatory else _OPTIONAL_FILL
        cell.font = _WHITE_BOLD
        cell.alignment = _CENTER
        ws1.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)
    ws1.row_dimensions[1].height = 22

    EXAMPLE_1 = [
        "TechCorp Systems", "PO-2026-001", "TechCorp Enterprise Suite FY26",
        "2026-04-01", "2027-03-31", "CLM-2026-001", "yes", "INR",
    ]
    ws1.append(EXAMPLE_1)
    for col_idx in range(1, len(EXAMPLE_1) + 1):
        ws1.cell(row=2, column=col_idx).font = _EXAMPLE_FONT
    ws1.row_dimensions[2].height = 18

    # Validations on row 3 (data row)
    dv_auto = DataValidation(type="list", formula1='"yes,no,opt_in"', allow_blank=True)
    dv_auto.sqref = "G3"
    ws1.add_data_validation(dv_auto)

    dv_currency = DataValidation(
        type="list",
        formula1='"INR,USD,EUR,GBP,JPY,CHF,AUD,CAD,SGD,AED"',
        allow_blank=True,
    )
    dv_currency.sqref = "H3"
    ws1.add_data_validation(dv_currency)

    dv_start = DataValidation(type="date", operator="greaterThan", formula1="DATE(2000,1,1)", allow_blank=True)
    dv_start.sqref = "D3"
    ws1.add_data_validation(dv_start)
    dv_end = DataValidation(type="date", operator="greaterThan", formula1="DATE(2000,1,1)", allow_blank=True)
    dv_end.sqref = "E3"
    ws1.add_data_validation(dv_end)

    # ── Tab 2: Contract Line Items ────────────────────────────────────────────
    ws2 = wb.create_sheet("Contract Line Items")

    # Row 1: section labels (merged bands)
    sections = [
        ("Step 3 — Software & Licensing", "A", "O"),
        ("Step 4 — Owner & DOA",          "P", "R"),
        ("Step 5 — Source & Usage Config","S", "T"),
    ]
    for label, col_from, col_to in sections:
        ws2.merge_cells(f"{col_from}1:{col_to}1")
        cell = ws2[f"{col_from}1"]
        cell.value = label
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        cell.alignment = _CENTER
    ws2.row_dimensions[1].height = 16

    # Row 2: column headers
    HEADERS_2 = [
        ("Software Name *",           True),   # A
        ("SW_ID",                     False),  # B
        ("License Type *",            True),   # C
        ("Metric *",                  True),   # D
        ("Entitled Count *",          True),   # E
        ("Unit Cost *",               True),   # F
        ("Annual Cost",               False),  # G
        ("Business Unit(s) *",        True),   # H
        ("Region(s) *",               True),   # I
        ("Category",                  False),  # J
        ("Sub-Category",              False),  # K
        ("GxP Flag",                  False),  # L
        ("Vendor Audit Risk",         False),  # M
        ("Deployment",                False),  # N
        ("Notes",                     False),  # O
        ("App Owner Email",           False),  # P
        ("Secondary Owner Email",     False),  # Q
        ("DOA Contact Email(s)",      False),  # R
        ("Discovery Source",          False),  # S
        ("Usage Update Method",       False),  # T
    ]
    for col_idx, (header, mandatory) in enumerate(HEADERS_2, 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.fill = _MANDATORY_FILL if mandatory else _OPTIONAL_FILL
        cell.font = _WHITE_BOLD
        cell.alignment = _CENTER
        width = max(len(header) + 4, 14)
        if header in ("Notes", "App Owner Email", "Secondary Owner Email", "DOA Contact Email(s)"):
            width = max(width, 30)
        ws2.column_dimensions[cell.column_letter].width = width
    ws2.row_dimensions[2].height = 22

    # Row 3: example
    EXAMPLE_2 = [
        "TechCorp ERP", "", "subscription", "Per User", 500, 8000, 4000000,
        "IT", "GG India", "ERP & Supply Chain", "ERP", "yes_21cfr", "MEDIUM", "cloud",
        "Core ERP module", "owner@drl.com", "", "", "SCCM (Microsoft MECM)",
        "Monthly Template Upload (XLSX)",
    ]
    ws2.append(EXAMPLE_2)
    for col_idx in range(1, len(EXAMPLE_2) + 1):
        ws2.cell(row=3, column=col_idx).font = _EXAMPLE_FONT
    ws2.row_dimensions[3].height = 18

    # Data validation on rows 4:1000
    def _list_dv(formula, col_letter):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showErrorMessage=True, errorTitle="Invalid",
                            error="Select a value from the dropdown list")
        dv.sqref = f"{col_letter}4:{col_letter}1000"
        return dv

    ws2.add_data_validation(_list_dv("'_Lists'!$A$2:$A$100", "C"))  # License Type
    ws2.add_data_validation(_list_dv("'_Lists'!$B$2:$B$100", "D"))  # Metric
    ws2.add_data_validation(_list_dv("'_Lists'!$C$2:$C$100", "H"))  # Business Units
    ws2.add_data_validation(_list_dv("'_Lists'!$D$2:$D$100", "I"))  # Regions
    ws2.add_data_validation(_list_dv("'_Lists'!$E$2:$E$100", "J"))  # Categories
    ws2.add_data_validation(_list_dv("'_Lists'!$F$2:$F$100", "K"))  # Sub-Categories
    ws2.add_data_validation(_list_dv("'_Lists'!$G$2:$G$100", "S"))  # Discovery Sources
    ws2.add_data_validation(_list_dv("'_Lists'!$H$2:$H$100", "T"))  # Usage Methods

    ws2.add_data_validation(DataValidation(
        type="list", formula1='"no,yes_21cfr,yes_annex11,yes_both"', allow_blank=True,
        sqref="L4:L1000"
    ))
    ws2.add_data_validation(DataValidation(
        type="list", formula1='"LOW,MEDIUM,HIGH"', allow_blank=True, sqref="M4:M1000"
    ))
    ws2.add_data_validation(DataValidation(
        type="list", formula1='"cloud,on_premise,desktop_cloud,hybrid"', allow_blank=True,
        sqref="N4:N1000"
    ))

    dv_int_e = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
    dv_int_e.sqref = "E4:E1000"
    ws2.add_data_validation(dv_int_e)
    dv_int_f = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_int_f.sqref = "F4:G1000"
    ws2.add_data_validation(dv_int_f)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_bulk_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    raw_headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(all_rows[0])]
    # Normalise headers to simple keys
    key_map = {
        "Software Name *": "primary_sw_name",
        "SW_ID (leave blank=new)": "sw_id",
        "Publisher": "publisher",
        "Category": "category_name",
        "Sub-Category": "sub_category_name",
        "Region": "region_name",
        "Deployment": "deployment",
        "GxP Relevant (yes/no)": "gxp",
        "Vendor Risk (LOW/MEDIUM/HIGH)": "vendor_risk",
        "Notes": "notes",
        "Contract Name *": "contract_name",
        "PO Number": "po_number",
        "CLM ID": "clm_id",
        "Start Date (YYYY-MM-DD)": "start_date",
        "End Date (YYYY-MM-DD)": "end_date",
        "Total Value (INR)": "total_value_inr",
        "Auto-Renewal (yes/no/opt_in)": "auto_renewal",
        "License Type (subscription/perpetual)": "license_type",
        "Metric": "metric",
        "Entitled Count": "entitled_count",
        "Unit Cost (INR)": "unit_cost_inr",
        "Annual Cost (INR)": "annual_cost_inr",
    }
    keys = [key_map.get(h, h.lower().replace(" ", "_")) for h in raw_headers]
    result = []
    for row in all_rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {keys[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(min(len(keys), len(row)))}
        if not d.get("primary_sw_name") or not d.get("contract_name"):
            continue  # skip rows missing required fields
        result.append(d)
    return result


def _parse_bulk_two_tab(data: bytes) -> tuple[dict, list[dict]]:
    """
    Parse a two-tab bulk onboarding XLSX.
    Tab 1 'Contract Information': row 1=headers, row 2=example, row 3=data.
    Tab 2 'Contract Line Items': row 1=section labels, row 2=headers, row 3=example, row 4+=data.
    Returns (contract_meta, line_items).
    Raises ValueError if the expected sheets are not found.
    """
    from openpyxl import load_workbook
    from datetime import datetime as _dt

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "Contract Information" not in wb.sheetnames:
        raise ValueError(
            "Sheet 'Contract Information' not found — please download the latest template"
        )
    if "Contract Line Items" not in wb.sheetnames:
        raise ValueError(
            "Sheet 'Contract Line Items' not found — please download the latest template"
        )

    # ── Tab 1: contract meta from row 3 ──────────────────────────────────────
    ws1 = wb["Contract Information"]
    rows1 = list(ws1.iter_rows(min_row=3, max_row=3, values_only=True))
    if not rows1 or not rows1[0] or all(v is None for v in rows1[0]):
        raise ValueError("Contract Information tab has no data in row 3")
    r = rows1[0]

    def _str(v) -> str | None:
        return str(v).strip() if v is not None and str(v).strip() else None

    def _date_str(v) -> str | None:
        if v is None:
            return None
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        try:
            _dt.strptime(s, "%Y-%m-%d")
            return s
        except ValueError:
            return None

    def _int(v) -> int | None:
        try:
            return int(v) if v is not None and str(v).strip() != "" else None
        except (ValueError, TypeError):
            return None

    def _split(v) -> list[str]:
        if not v or str(v).strip() == "":
            return []
        return [x.strip() for x in str(v).split(",") if x.strip()]

    auto_renewal_raw = _str(r[6])
    auto_renewal = auto_renewal_raw if auto_renewal_raw in ("yes", "no", "opt_in") else None
    currency = _str(r[7]) or "INR"

    contract_meta = {
        "vendor_name": _str(r[0]),
        "po_number": _str(r[1]),
        "contract_name": _str(r[2]),
        "start_date": _date_str(r[3]),
        "end_date": _date_str(r[4]),
        "clm_id": _str(r[5]),
        "auto_renewal_clause": auto_renewal,
        "currency": currency,
    }

    # ── Tab 2: line items from row 4 onward ───────────────────────────────────
    ws2 = wb["Contract Line Items"]
    all_rows = list(ws2.iter_rows(min_row=4, values_only=True))
    line_items = []
    for row in all_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        primary_sw_name = _str(row[0])
        if not primary_sw_name:
            continue  # skip rows without software name

        gxp_raw = _str(row[11]) or "no"
        gxp_flag = gxp_raw if gxp_raw in ("no", "yes_21cfr", "yes_annex11", "yes_both") else "no"
        risk_raw = (_str(row[12]) or "LOW").upper()
        vendor_risk = risk_raw if risk_raw in ("LOW", "MEDIUM", "HIGH") else "LOW"
        deploy_raw = _str(row[13]) or "cloud"
        deployment = deploy_raw if deploy_raw in ("cloud", "on_premise", "desktop_cloud", "hybrid") else "cloud"

        line_items.append({
            "primary_sw_name": primary_sw_name,
            "sw_id": _str(row[1]),
            "license_type_name": _str(row[2]),
            "metric_name": _str(row[3]),
            "entitled_count": _int(row[4]),
            "unit_cost": _int(row[5]),
            "annual_cost": _int(row[6]),
            "business_units": _split(row[7]),
            "regions": _split(row[8]),
            "category_name": _str(row[9]),
            "sub_category_name": _str(row[10]),
            "gxp_flag": gxp_flag,
            "vendor_risk": vendor_risk,
            "deployment": deployment,
            "notes": _str(row[14]),
            "app_owner_email": _str(row[15]),
            "secondary_owner_email": _str(row[16]),
            "doa_emails": _split(row[17]),
            "discovery_source_name": _str(row[18]),
            "usage_method_name": _str(row[19]),
        })

    return contract_meta, line_items


@router.get("/bulk-template")
async def download_bulk_template(db: AsyncSession = Depends(get_db)):
    from app.models.masters import LicenseType, LicenseMetric, Region, Category, SubCategory, DiscoverySource, UsageUpdateMethod

    def _names(rows):
        return [r.name for r in rows]

    license_types = [r.license_type for r in (await db.execute(select(LicenseType))).scalars().all()]
    metrics       = _names((await db.execute(select(LicenseMetric))).scalars().all())
    regions       = _names((await db.execute(select(Region))).scalars().all())
    categories    = _names((await db.execute(select(Category))).scalars().all())
    sub_categories = _names((await db.execute(select(SubCategory))).scalars().all())
    discovery_sources = _names((await db.execute(select(DiscoverySource))).scalars().all())
    usage_methods = _names((await db.execute(select(UsageUpdateMethod))).scalars().all())

    db_lists = {
        "license_types": license_types,
        "metrics": metrics,
        "regions": regions,
        "categories": categories,
        "sub_categories": sub_categories,
        "discovery_sources": discovery_sources,
        "usage_methods": usage_methods,
    }
    data = _generate_bulk_template(db_lists)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=DRL_BulkOnboarding_Template.xlsx"},
    )


@router.post("/bulk", status_code=201)
async def bulk_onboard(
    file: UploadFile = File(...),
    current_user=Depends(require_role(["COE_ADMIN"])),
    db: AsyncSession = Depends(get_db),
):
    """
    Process bulk onboarding XLSX. Each row creates a SoftwareCatalog entry
    (or maps to existing), a Contract, and an Entitlement. Returns a summary.
    """
    from datetime import datetime as _dt
    from app.models.masters import Category, LicenseMetric

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    data = await file.read()
    rows = _parse_bulk_xlsx(data)
    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found. Check that Software Name and Contract Name columns are filled.")

    created_sw: list[str] = []
    created_ent: list[str] = []
    skipped: list[str] = []

    for idx, row in enumerate(rows, 1):
        try:
            canonical = row["primary_sw_name"]
            contract_name = row["contract_name"]

            # ── Resolve SW_ID ──────────────────────────────────────────────
            sw_id_input = row.get("sw_id", "").strip()
            if sw_id_input:
                sw = await db.get(SoftwareCatalog, sw_id_input)
                if not sw:
                    skipped.append(f"Row {idx}: SW_ID '{sw_id_input}' not found — skipped")
                    continue
                sw_id = sw.sw_id
            else:
                # Check if canonical name already exists
                existing = (await db.execute(
                    select(SoftwareCatalog).where(SoftwareCatalog.primary_sw_name == canonical)
                )).scalar_one_or_none()
                if existing:
                    sw_id = existing.sw_id
                else:
                    sw_id = await _next_sw_id(db, row.get("publisher") or None)
                    gxp_raw = row.get("gxp", "no").lower()
                    gxp_flag = "yes_21cfr" if gxp_raw == "yes" else "no"
                    vendor_risk = row.get("vendor_risk", "LOW").upper()
                    if vendor_risk not in ("LOW", "MEDIUM", "HIGH"):
                        vendor_risk = "LOW"
                    deploy = row.get("deployment", "cloud").lower().replace(" ", "_").replace("/", "_")
                    if deploy not in ("cloud", "on_premise", "desktop_cloud", "hybrid"):
                        deploy = "cloud"

                    # Resolve category + sub-category by name
                    from app.models.masters import SubCategory, Region
                    cat_name = row.get("category_name", "").strip()
                    cat_id = None
                    sub_cat_id = None
                    if cat_name:
                        cat = (await db.execute(
                            select(Category).where(Category.name.ilike(cat_name))
                        )).scalar_one_or_none()
                        if cat:
                            cat_id = cat.id
                            sub_cat_name = row.get("sub_category_name", "").strip()
                            if sub_cat_name:
                                sub = (await db.execute(
                                    select(SubCategory).where(
                                        SubCategory.category_id == cat.id,
                                        SubCategory.name.ilike(sub_cat_name)
                                    )
                                )).scalar_one_or_none()
                                if sub:
                                    sub_cat_id = sub.id

                    # Resolve region by name
                    region_id = None
                    region_name_val = row.get("region_name", "").strip()
                    if region_name_val:
                        reg = (await db.execute(
                            select(Region).where(Region.name.ilike(region_name_val))
                        )).scalar_one_or_none()
                        if reg:
                            region_id = reg.id

                    sw_entry = SoftwareCatalog(
                        sw_id=sw_id,
                        primary_sw_name=canonical,
                        publisher=row.get("publisher") or None,
                        category_id=cat_id,
                        sub_category_id=sub_cat_id,
                        gxp_flag=gxp_flag,
                        vendor_risk=vendor_risk,
                        deployment=deploy,
                        region_id=region_id,
                        notes=row.get("notes") or None,
                        onboarded_date=date.today(),
                    )
                    db.add(sw_entry)
                    await db.flush()
                    created_sw.append(sw_id)

            # ── Parse dates ────────────────────────────────────────────────
            def _d(s):
                if not s:
                    return None
                try:
                    return _dt.strptime(s.strip(), "%Y-%m-%d").date()
                except (ValueError, AttributeError):
                    return None

            # ── Create Contract ────────────────────────────────────────────
            auto_r = row.get("auto_renewal", "").lower()
            if auto_r not in ("yes", "no", "opt_in"):
                auto_r = None

            contract = Contract(
                sw_id=sw_id,
                po_number=row.get("po_number") or None,
                clm_id=row.get("clm_id") or None,
                start_date=_d(row.get("start_date")),
                end_date=_d(row.get("end_date")),
                total_value_inr=int(row["total_value_inr"]) if row.get("total_value_inr") else None,
                auto_renewal_clause=auto_r,
                storage_backend="local",
                created_by=current_user.id,
            )
            db.add(contract)
            await db.flush()

            # ── Resolve metric ─────────────────────────────────────────────
            metric_name = row.get("metric", "").strip()
            metric_id = None
            if metric_name:
                metric = (await db.execute(
                    select(LicenseMetric).where(LicenseMetric.name.ilike(metric_name))
                )).scalar_one_or_none()
                if metric:
                    metric_id = metric.id

            # ── Create Entitlement ─────────────────────────────────────────
            max_ent = (await db.execute(
                select(func.max(Entitlement.ent_id)).where(Entitlement.ent_id.like("ENT-%"))
            )).scalar_one_or_none()
            ent_n = int(max_ent.split("-")[1]) + 1 if max_ent else 1
            while True:
                ent_candidate = f"ENT-{ent_n:03d}"
                existing_ent = await db.get(Entitlement, ent_candidate)
                if not existing_ent:
                    ent_id = ent_candidate
                    break
                ent_n += 1

            # Resolve license type from license_types table
            from app.models.masters import LicenseType
            lic_type_str = row.get("license_type", "subscription").lower()
            lt_row = (await db.execute(
                select(LicenseType).where(LicenseType.license_type == lic_type_str)
            )).scalar_one_or_none()
            license_type_id_val = lt_row.id if lt_row else None

            ent = Entitlement(
                ent_id=ent_id,
                sw_id=sw_id,
                contract_id=contract.id,
                contract_name=contract_name,
                metric_id=metric_id,
                license_type_id=license_type_id_val,
                entitled_count=int(row["entitled_count"]) if row.get("entitled_count") else None,
                in_use_count=0,
                unit_cost=int(row["unit_cost_inr"]) if row.get("unit_cost_inr") else None,
                annual_cost=int(row["annual_cost_inr"]) if row.get("annual_cost_inr") else None,
                notes=row.get("notes") or None,
                status="ACTIVE",
            )
            db.add(ent)
            await db.flush()
            created_ent.append(ent_id)

        except Exception as e:
            skipped.append(f"Row {idx} ({row.get('primary_sw_name','?')}): {e}")

    await db.commit()
    return {
        "sw_created": len(created_sw),
        "ent_created": len(created_ent),
        "sw_ids": created_sw,
        "ent_ids": created_ent,
        "skipped": skipped,
    }
