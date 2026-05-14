import io
import uuid
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.discovery import DiscoveryRecord
from app.models.catalog import SoftwareCatalog, SoftwareAlias
from app.schemas.discovery import DiscoveryRecordOut, IngestResultOut

router = APIRouter(prefix="/discovery", tags=["discovery"])


async def _next_disc_ids(db: AsyncSession, count: int) -> list[str]:
    """Returns `count` sequential disc_ids starting from current max+1."""
    result = await db.execute(
        select(func.max(DiscoveryRecord.disc_id)).where(DiscoveryRecord.disc_id.like("D-%"))
    )
    max_id = result.scalar_one_or_none()
    n = int(max_id.split("-")[1]) + 1 if max_id else 1
    return [f"D-{(n + i):04d}" for i in range(count)]


async def _resolve_sw_id(contract_name: str, db: AsyncSession) -> str | None:
    """Try to match contract_name to SW catalog via canonical_name or alias."""
    result = await db.execute(
        select(SoftwareCatalog.sw_id).where(
            SoftwareCatalog.canonical_name.ilike(contract_name)
        )
    )
    sw_id = result.scalar_one_or_none()
    if sw_id:
        return sw_id
    result = await db.execute(
        select(SoftwareAlias.sw_id).where(
            SoftwareAlias.alias_name.ilike(contract_name)
        )
    )
    return result.scalar_one_or_none()


def _parse_csv_or_xlsx(data: bytes, filename: str) -> list[dict]:
    """
    Parse discovery upload CSV or XLSX.
    Expected columns: contract_name | device_id | device_type | os | version | last_seen | site
    """
    if filename.lower().endswith(".csv"):
        import csv
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
        return [dict(zip(headers, row)) for row in rows[1:] if any(row)]


@router.get("", response_model=list[DiscoveryRecordOut])
async def list_discovery(
    sw_id: str | None = Query(None),
    matched: bool | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    q = select(DiscoveryRecord)
    if sw_id:
        q = q.where(DiscoveryRecord.sw_id == sw_id)
    if matched is True:
        q = q.where(DiscoveryRecord.sw_id.is_not(None))
    elif matched is False:
        q = q.where(DiscoveryRecord.sw_id.is_(None))
    result = await db.execute(q.order_by(DiscoveryRecord.disc_id))
    return [DiscoveryRecordOut.model_validate(r) for r in result.scalars().all()]


@router.post("/ingest", response_model=IngestResultOut, status_code=201)
async def ingest_discovery(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    lower = file.filename.lower()
    if not (lower.endswith(".csv") or lower.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Upload must be .csv or .xlsx")

    data = await file.read()
    rows = _parse_csv_or_xlsx(data, file.filename)
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    batch_id = uuid.uuid4()
    today = date.today()
    errors: list[str] = []
    valid_rows = [r for r in rows if (r.get("contract_name") or r.get("Contract Name") or r.get("Contract_Name") or "").strip()]
    skipped = len(rows) - len(valid_rows)
    if skipped:
        errors.append(f"{skipped} rows skipped (missing contract_name)")

    disc_ids = await _next_disc_ids(db, len(valid_rows))
    inserted = 0
    matched = 0

    for i, row in enumerate(valid_rows):
        contract_name = (row.get("contract_name") or row.get("Contract Name") or row.get("Contract_Name") or "").strip()

        # Parse last_seen date
        raw_last_seen = row.get("last_seen") or row.get("Last Seen") or row.get("Last_Seen")
        last_seen_date = None
        if raw_last_seen:
            try:
                if isinstance(raw_last_seen, (date, datetime)):
                    last_seen_date = raw_last_seen.date() if isinstance(raw_last_seen, datetime) else raw_last_seen
                else:
                    last_seen_date = datetime.strptime(str(raw_last_seen).strip(), "%Y-%m-%d").date()
            except ValueError:
                errors.append(f"Row {i+2}: invalid last_seen '{raw_last_seen}' — stored as null")

        device_type_raw = (row.get("device_type") or row.get("Device Type") or "endpoint").strip().lower()
        device_type = device_type_raw if device_type_raw in ("endpoint", "server") else "endpoint"

        sw_id = await _resolve_sw_id(contract_name, db)
        if sw_id:
            matched += 1

        rec = DiscoveryRecord(
            disc_id=disc_ids[inserted],
            contract_name=contract_name,
            sw_id=sw_id,
            device_id=(row.get("device_id") or row.get("Device ID") or "").strip() or None,
            device_type=device_type,
            os=(row.get("os") or row.get("OS") or "").strip() or None,
            version=(row.get("version") or row.get("Version") or "").strip() or None,
            last_seen=last_seen_date,
            site=(row.get("site") or row.get("Site") or "").strip() or None,
            upload_date=today,
            upload_batch_id=batch_id,
        )
        db.add(rec)
        inserted += 1

    await db.commit()
    return IngestResultOut(
        batch_id=batch_id,
        inserted=inserted,
        matched=matched,
        unmatched=inserted - matched,
        errors=errors,
    )
