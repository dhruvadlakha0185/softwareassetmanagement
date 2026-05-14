import io
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.audit import AuditTrail
from app.schemas.audit import AuditTrailOut

router = APIRouter(prefix="/audit", tags=["audit"])


@router.get("", response_model=list[AuditTrailOut])
async def list_audit(
    entity_type: str | None = Query(None),
    action_type: str | None = Query(None),
    sw_id: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    limit: int = Query(100, le=500),
    offset: int = Query(0),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(AuditTrail).order_by(AuditTrail.created_at_utc.desc())
    if entity_type:
        q = q.where(AuditTrail.entity_type == entity_type)
    if action_type:
        q = q.where(AuditTrail.action_type == action_type)
    if sw_id:
        q = q.where(AuditTrail.sw_id == sw_id)
    if date_from:
        q = q.where(AuditTrail.created_at_utc >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(AuditTrail.created_at_utc <= datetime.combine(date_to, datetime.max.time()))
    q = q.limit(limit).offset(offset)
    result = await db.execute(q)
    return [AuditTrailOut.model_validate(r) for r in result.scalars().all()]


@router.get("/export")
async def export_audit(
    entity_type: str | None = Query(None),
    sw_id: str | None = Query(None),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(AuditTrail).order_by(AuditTrail.created_at_utc.desc()).limit(1000)
    if entity_type:
        q = q.where(AuditTrail.entity_type == entity_type)
    if sw_id:
        q = q.where(AuditTrail.sw_id == sw_id)
    result = await db.execute(q)
    rows = result.scalars().all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    headers = ["Timestamp (UTC)", "Action", "Entity Type", "Entity ID",
               "SW_ID", "User ID", "GxP", "Reason for Change"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A2E5A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            str(r.created_at_utc) if r.created_at_utc else "",
            r.action_type,
            r.entity_type,
            r.entity_id or "",
            r.sw_id or "",
            str(r.user_id) if r.user_id else "",
            "YES" if r.is_gxp else "NO",
            r.reason_for_change or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    today = date.today().isoformat()
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=audit_trail_{today}.xlsx"},
    )
