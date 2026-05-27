import io
import pytest
from unittest.mock import AsyncMock, patch
from openpyxl import Workbook
from app.models.contracts import EntitlementDoaContact
from app.schemas.onboarding import MultiLineItemIn, MultiPublishPayload

MOCK_EXTRACTION = {
    "vendor_name": "Microsoft Corporation",
    "po_number": "PO-2024-001",
    "clm_id": None,
    "start_date": "2024-04-01",
    "end_date": "2025-03-31",
    "auto_renewal_clause": "yes",
    "total_value_inr": 5000000,
    "reseller": None,
    "line_items": [
        {
            "contract_name": "Microsoft 365 E3",
            "metric": "Per User",
            "license_type": "subscription",
            "entitled_count": 500,
            "unit_cost_inr": 10000,
            "annual_cost_inr": 5000000,
        }
    ],
}


async def test_extract_requires_auth(client):
    resp = await client.post(
        "/api/v1/onboarding/extract",
        files={"file": ("test.pdf", io.BytesIO(b"%PDF-1.4 test"), "application/pdf")},
    )
    assert resp.status_code in (401, 403)


async def test_extract_returns_json(client, admin_token):
    h = {"Authorization": f"Bearer {admin_token}"}
    with patch("app.api.v1.routes.onboarding.call_openai", new_callable=AsyncMock) as mock_ai, \
         patch("app.api.v1.routes.onboarding.extract_contract_text", return_value="contract text"):
        mock_ai.return_value = MOCK_EXTRACTION
        resp = await client.post(
            "/api/v1/onboarding/extract",
            files={"file": ("contract.pdf", io.BytesIO(b"%PDF-1.4"), "application/pdf")},
            headers=h,
        )
    assert resp.status_code == 200
    data = resp.json()
    assert data["po_number"] == "PO-2024-001"
    assert len(data["line_items"]) == 1


async def test_draft_lifecycle(client, admin_token):
    h = {"Authorization": f"Bearer {admin_token}"}

    # create
    create = await client.post(
        "/api/v1/onboarding/drafts",
        json={"po_number": "PO-DRAFT-01", "current_step": 2},
        headers=h,
    )
    assert create.status_code == 201
    draft_id = create.json()["id"]

    # get
    get_resp = await client.get(f"/api/v1/onboarding/drafts/{draft_id}", headers=h)
    assert get_resp.status_code == 200
    assert get_resp.json()["po_number"] == "PO-DRAFT-01"

    # update
    update = await client.put(
        f"/api/v1/onboarding/drafts/{draft_id}",
        json={"current_step": 4},
        headers=h,
    )
    assert update.status_code == 200
    assert update.json()["current_step"] == 4

    # list
    drafts = (await client.get("/api/v1/onboarding/drafts", headers=h)).json()
    assert any(d["id"] == draft_id for d in drafts)

    # delete
    del_resp = await client.delete(f"/api/v1/onboarding/drafts/{draft_id}", headers=h)
    assert del_resp.status_code == 204


async def test_publish_creates_sw_contract_entitlement(client, admin_token):
    h = {"Authorization": f"Bearer {admin_token}"}
    payload = {
        "primary_sw_name": "Publish Test Software",
        "publisher": "Pub Corp",
        "gxp_flag": "no",
        "vendor_risk": "LOW",
        "deployment": "cloud",
        "po_number": "PO-PUB-001",
        "start_date": "2024-01-01",
        "end_date": "2025-01-01",
        "line_items": [
            {
                "contract_name": "Publish Test E3",
                "metric": "Per User",
                "license_type": "subscription",
                "entitled_count": 100,
                "unit_cost_inr": 1000,
                "annual_cost_inr": 100000,
            }
        ],
        "aliases": ["PubTest", "PT Software"],
    }
    resp = await client.post("/api/v1/onboarding/publish", json=payload, headers=h)
    assert resp.status_code == 201
    data = resp.json()
    assert data["sw_id"].startswith("SW-")
    assert len(data["ent_ids"]) == 1
    assert data["ent_ids"][0].startswith("ENT-")

    # Verify SW entry exists with aliases
    catalog_entry = (await client.get(f"/api/v1/catalog/{data['sw_id']}")).json()
    assert catalog_entry["primary_sw_name"] == "Publish Test Software"
    alias_names = [a["alias_name"] for a in catalog_entry["aliases"]]
    assert "PubTest" in alias_names
    # No cleanup needed — contracts FK prevents catalog delete; test DB drops all at session end


async def test_publish_maps_to_existing_sw(client, admin_token):
    h = {"Authorization": f"Bearer {admin_token}"}
    # create SW first
    sw = (await client.post("/api/v1/catalog", json={
        "primary_sw_name": "Existing SW For Publish",
        "gxp_flag": "no", "vendor_risk": "LOW", "deployment": "cloud",
    }, headers=h)).json()
    sw_id = sw["sw_id"]

    payload = {
        "primary_sw_name": "Existing SW For Publish",
        "sw_id": sw_id,
        "line_items": [
            {"contract_name": "Existing SW Sub", "license_type": "subscription", "entitled_count": 50}
        ],
    }
    resp = await client.post("/api/v1/onboarding/publish", json=payload, headers=h)
    assert resp.status_code == 201
    assert resp.json()["sw_id"] == sw_id
    # No cleanup — contracts FK prevents catalog delete; test DB drops all at session end


def test_entitlement_doa_contact_model_has_expected_columns():
    cols = {c.key for c in EntitlementDoaContact.__table__.columns}
    assert {"id", "ent_id", "doa_contact_id"}.issubset(cols)


def test_multi_line_item_in_has_per_item_owner_fields():
    item = MultiLineItemIn(contract_name="Test", primary_sw_name="TestSW")
    assert hasattr(item, "app_owner_id")
    assert hasattr(item, "secondary_owner_id")
    assert hasattr(item, "doa_contact_ids")
    assert hasattr(item, "discovery_source_id")
    assert hasattr(item, "usage_method_id")
    # defaults
    assert item.doa_contact_ids == []
    assert item.app_owner_id is None


def test_multi_publish_payload_no_longer_has_owner_fields():
    payload = MultiPublishPayload(
        vendor_name="ACME",
        line_items=[],
    )
    assert not hasattr(payload, "app_owner_id")
    assert not hasattr(payload, "secondary_owner_id")
    assert not hasattr(payload, "discovery_source_id")
    assert not hasattr(payload, "usage_method_id")


@pytest.mark.asyncio
async def test_multi_publish_writes_per_item_owner_fields(client, admin_token, db):
    """owner fields on the line item are persisted to the entitlement."""
    h = {"Authorization": f"Bearer {admin_token}"}

    # Get master data IDs from /api/v1/masters/all
    masters_resp = await client.get("/api/v1/masters/all")
    assert masters_resp.status_code == 200
    masters = masters_resp.json()

    sources = masters.get("sources", [])
    methods = masters.get("methods", [])

    source_id = sources[0]["id"] if sources else None
    method_id = methods[0]["id"] if methods else None

    payload = {
        "vendor_name": "TestCo",
        "po_number": "PO-TEST-OWNER-001",
        "line_items": [
            {
                "contract_name": "TestSW Owner License",
                "primary_sw_name": "TestSWOwner",
                "discovery_source_id": source_id,
                "usage_method_id": method_id,
                "deployment": "cloud",
                "gxp_flag": "no",
            }
        ],
    }
    resp = await client.post("/api/v1/onboarding/multi-publish", json=payload, headers=h)
    assert resp.status_code == 201, resp.text
    result = resp.json()
    assert len(result["created"]) == 1
    ent_id = result["created"][0]["ent_id"]

    from sqlalchemy import select
    from app.models.contracts import Entitlement
    result_ent = await db.execute(select(Entitlement).where(Entitlement.ent_id == ent_id))
    ent = result_ent.scalar_one()
    if source_id:
        assert str(ent.discovery_source_id) == source_id
    if method_id:
        assert str(ent.usage_method_id) == method_id


@pytest.mark.asyncio
async def test_multi_publish_creates_doa_contact_rows(client, admin_token, db):
    """doa_contact_ids on the line item are persisted to entitlement_doa_contacts."""
    from sqlalchemy import select, text
    from app.models.contracts import EntitlementDoaContact
    from app.models.users import DOAHierarchy
    import uuid

    # Get or create a doa_hierarchy row
    result = await db.execute(text("SELECT id FROM doa_hierarchy LIMIT 1"))
    row = result.first()

    if row is None:
        doa_id = str(uuid.uuid4())
        await db.execute(text(
            "INSERT INTO doa_hierarchy (id, full_name, email) "
            "VALUES (:id, :full_name, :email)"
        ), {"id": doa_id, "full_name": "Test DOA Contact", "email": "doa@test.com"})
        await db.flush()
    else:
        doa_id = str(row[0])

    h = {"Authorization": f"Bearer {admin_token}"}
    payload = {
        "vendor_name": "TestCo",
        "po_number": "PO-DOA-TEST-001",
        "line_items": [
            {
                "contract_name": "DOA Test License",
                "primary_sw_name": "DOATestSW",
                "doa_contact_ids": [doa_id],
                "deployment": "cloud",
                "gxp_flag": "no",
            }
        ],
    }
    resp = await client.post("/api/v1/onboarding/multi-publish", json=payload, headers=h)
    assert resp.status_code == 201, resp.text
    result_data = resp.json()
    assert len(result_data["created"]) == 1, (
        f"Expected 1 created, got skipped: {result_data.get('skipped')}"
    )
    ent_id = result_data["created"][0]["ent_id"]

    result = await db.execute(
        select(EntitlementDoaContact).where(EntitlementDoaContact.ent_id == ent_id)
    )
    rows = result.scalars().all()
    assert len(rows) == 1
    assert str(rows[0].doa_contact_id) == doa_id


@pytest.mark.asyncio
async def test_multi_publish_generates_notes_when_blank(client, admin_token, db):
    """When notes is None on the line item, the generator is called and the result is persisted."""
    from sqlalchemy import select
    from app.models.contracts import Entitlement
    from unittest.mock import AsyncMock, patch

    h = {"Authorization": f"Bearer {admin_token}"}
    payload = {
        "vendor_name": "TestVendor",
        "po_number": "PO-NOTES-GEN-001",
        "line_items": [
            {
                "contract_name": "TestSW Notes License",
                "primary_sw_name": "TestSWNotes",
                "deployment": "cloud",
                "gxp_flag": "no",
            }
        ],
    }

    with patch(
        "app.api.v1.routes.onboarding.generate_entitlement_notes",
        new_callable=AsyncMock,
        return_value="Generated procurement note for testing.",
    ) as mock_gen:
        resp = await client.post("/api/v1/onboarding/multi-publish", json=payload, headers=h)

    assert resp.status_code == 201, resp.text
    ent_id = resp.json()["created"][0]["ent_id"]
    mock_gen.assert_called_once()

    result = await db.execute(select(Entitlement).where(Entitlement.ent_id == ent_id))
    ent = result.scalar_one()
    assert ent.notes == "Generated procurement note for testing."


@pytest.mark.asyncio
async def test_multi_publish_skips_generator_when_notes_present(client, admin_token, db):
    """When the line item already has notes, the generator is NOT called and existing notes persist."""
    from sqlalchemy import select
    from app.models.contracts import Entitlement
    from unittest.mock import AsyncMock, patch

    h = {"Authorization": f"Bearer {admin_token}"}
    payload = {
        "vendor_name": "TestVendor",
        "po_number": "PO-NOTES-SKIP-001",
        "line_items": [
            {
                "contract_name": "TestSW Existing Notes",
                "primary_sw_name": "TestSWExistingNotes",
                "notes": "Existing handwritten note.",
                "deployment": "cloud",
                "gxp_flag": "no",
            }
        ],
    }

    with patch(
        "app.api.v1.routes.onboarding.generate_entitlement_notes",
        new_callable=AsyncMock,
        return_value="Should not appear.",
    ) as mock_gen:
        resp = await client.post("/api/v1/onboarding/multi-publish", json=payload, headers=h)

    assert resp.status_code == 201, resp.text
    mock_gen.assert_not_called()

    ent_id = resp.json()["created"][0]["ent_id"]
    result = await db.execute(select(Entitlement).where(Entitlement.ent_id == ent_id))
    ent = result.scalar_one()
    assert ent.notes == "Existing handwritten note."


def _make_two_tab_xlsx(
    meta_row=None,
    item_rows=None,
) -> bytes:
    """Helper: build a minimal valid two-tab XLSX for parser tests."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Contract Information"
    # Row 1: headers (parser reads row 3 for data)
    ws1.append(["Vendor / Publisher Name *", "PO Number *", "Contract Name *",
                "Contract Start Date *", "Contract End Date *", "CLM ID",
                "Auto-Renewal Clause", "Currency"])
    # Row 2: example
    ws1.append(["Example Corp", "PO-EX-001", "Example Contract", "2026-01-01", "2027-01-01", "", "yes", "INR"])
    # Row 3: data
    row = meta_row or ["TechVendor", "PO-2026-001", "TV Contract FY26",
                       "2026-04-01", "2027-03-31", "CLM-001", "yes", "USD"]
    ws1.append(row)

    ws2 = wb.create_sheet("Contract Line Items")
    # Row 1: section labels
    ws2.append(["Step 3"] + [""] * 14 + ["Step 4"] + [""] * 2 + ["Step 5"] + [""])
    # Row 2: headers
    ws2.append(["Software Name *", "SW_ID", "License Type *", "Metric *",
                "Entitled Count *", "Unit Cost *", "Annual Cost",
                "Business Unit(s) *", "Region(s) *", "Category", "Sub-Category",
                "GxP Flag", "Vendor Audit Risk", "Deployment", "Notes",
                "App Owner Email", "Secondary Owner Email", "DOA Contact Email(s)",
                "Discovery Source", "Usage Update Method"])
    # Row 3: example
    ws2.append(["ExSW", "", "subscription", "Per User", 100, 5000, 500000,
                "IT", "Global", "ERP & Supply Chain", "ERP", "no", "LOW", "cloud", "Example notes",
                "ex@drl.com", "", "", "SCCM (Microsoft MECM)", "Monthly Template Upload (XLSX)"])
    # Row 4+: data
    for row in (item_rows or [
        ["My Software", "SW-001", "subscription", "Per User",
         250, 8000, None,
         "IT, Finance", "GG India, EUG", "Enterprise Productivity", "Low-Code",
         "yes_21cfr", "MEDIUM", "cloud", "Core tool",
         "owner@drl.com", "", "doa@drl.com",
         "ServiceNow CMDB", "Quarterly Manual Update"]
    ]):
        ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_two_tab_contract_meta():
    from app.api.v1.routes.onboarding import _parse_bulk_two_tab
    meta, _ = _parse_bulk_two_tab(_make_two_tab_xlsx())
    assert meta["vendor_name"] == "TechVendor"
    assert meta["po_number"] == "PO-2026-001"
    assert meta["contract_name"] == "TV Contract FY26"
    assert meta["start_date"] == "2026-04-01"
    assert meta["end_date"] == "2027-03-31"
    assert meta["clm_id"] == "CLM-001"
    assert meta["auto_renewal_clause"] == "yes"
    assert meta["currency"] == "USD"


def test_parse_two_tab_line_items():
    from app.api.v1.routes.onboarding import _parse_bulk_two_tab
    _, items = _parse_bulk_two_tab(_make_two_tab_xlsx())
    assert len(items) == 1
    item = items[0]
    assert item["primary_sw_name"] == "My Software"
    assert item["sw_id"] == "SW-001"
    assert item["license_type_name"] == "subscription"
    assert item["metric_name"] == "Per User"
    assert item["entitled_count"] == 250
    assert item["unit_cost"] == 8000
    assert item["annual_cost"] is None  # blank → None (auto-calc done later)
    assert item["business_units"] == ["IT", "Finance"]
    assert item["regions"] == ["GG India", "EUG"]
    assert item["category_name"] == "Enterprise Productivity"
    assert item["sub_category_name"] == "Low-Code"
    assert item["gxp_flag"] == "yes_21cfr"
    assert item["vendor_risk"] == "MEDIUM"
    assert item["deployment"] == "cloud"
    assert item["notes"] == "Core tool"
    assert item["app_owner_email"] == "owner@drl.com"
    assert item["secondary_owner_email"] is None
    assert item["doa_emails"] == ["doa@drl.com"]
    assert item["discovery_source_name"] == "ServiceNow CMDB"
    assert item["usage_method_name"] == "Quarterly Manual Update"


def test_parse_two_tab_skips_blank_rows():
    from app.api.v1.routes.onboarding import _parse_bulk_two_tab
    xlsx = _make_two_tab_xlsx(item_rows=[
        ["Real Software", "", "subscription", "Per User", 10, 1000, None, "IT", "Global",
         "", "", "no", "LOW", "cloud", "", "", "", "", "", ""],
        [None, None, None, None, None, None, None, None, None,
         None, None, None, None, None, None, None, None, None, None, None],
    ])
    _, items = _parse_bulk_two_tab(xlsx)
    assert len(items) == 1
    assert items[0]["primary_sw_name"] == "Real Software"


def test_parse_two_tab_old_format_raises():
    from app.api.v1.routes.onboarding import _parse_bulk_two_tab
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Onboarding"  # old single-sheet format
    ws.append(["Software Name *", "Contract Name *"])
    ws.append(["SAP ERP", "SAP FY26"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Contract Information"):
        _parse_bulk_two_tab(buf.getvalue())


def test_generate_template_has_correct_sheets():
    from app.api.v1.routes.onboarding import _generate_bulk_template
    from openpyxl import load_workbook
    db_lists = {
        "license_types": ["perpetual", "subscription"],
        "metrics": ["Per User", "Per Core (2-pack)"],
        "regions": ["GG India", "Global", "EUG"],
        "categories": ["ERP & Supply Chain", "IT Security"],
        "sub_categories": ["ERP", "EDR/XDR"],
        "discovery_sources": ["SCCM (Microsoft MECM)", "ServiceNow CMDB"],
        "usage_methods": ["Monthly Template Upload (XLSX)", "Quarterly Manual Update"],
    }
    data = _generate_bulk_template(db_lists)
    wb = load_workbook(io.BytesIO(data))
    assert "Contract Information" in wb.sheetnames
    assert "Contract Line Items" in wb.sheetnames
    assert "_Lists" in wb.sheetnames


def test_generate_template_lists_sheet_is_hidden():
    from app.api.v1.routes.onboarding import _generate_bulk_template
    from openpyxl import load_workbook
    db_lists = {
        "license_types": ["subscription"],
        "metrics": ["Per User"],
        "regions": ["Global"],
        "categories": ["IT Security"],
        "sub_categories": ["Firewall"],
        "discovery_sources": ["Manual / Procurement"],
        "usage_methods": ["Quarterly Manual Update"],
    }
    data = _generate_bulk_template(db_lists)
    wb = load_workbook(io.BytesIO(data))
    assert wb["_Lists"].sheet_state == "hidden"


def test_generate_template_lists_sheet_contains_db_values():
    from app.api.v1.routes.onboarding import _generate_bulk_template
    from openpyxl import load_workbook
    db_lists = {
        "license_types": ["perpetual", "subscription"],
        "metrics": ["Per User"],
        "regions": ["GG India", "Global"],
        "categories": ["IT Security"],
        "sub_categories": ["Firewall"],
        "discovery_sources": ["ServiceNow CMDB"],
        "usage_methods": ["Quarterly Manual Update"],
    }
    data = _generate_bulk_template(db_lists)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["_Lists"]
    # Col A = license_types (header row 1, values from row 2)
    col_a = [ws.cell(row=r, column=1).value for r in range(2, 10) if ws.cell(row=r, column=1).value]
    assert "perpetual" in col_a
    assert "subscription" in col_a
    # Col D = regions
    col_d = [ws.cell(row=r, column=4).value for r in range(2, 10) if ws.cell(row=r, column=4).value]
    assert "GG India" in col_d
    assert "Global" in col_d


def test_generate_template_contract_info_has_8_columns():
    from app.api.v1.routes.onboarding import _generate_bulk_template
    from openpyxl import load_workbook
    db_lists = {
        "license_types": [], "metrics": [], "regions": [],
        "categories": [], "sub_categories": [], "discovery_sources": [], "usage_methods": [],
    }
    data = _generate_bulk_template(db_lists)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Contract Information"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert headers[0] == "Vendor / Publisher Name *"
    assert headers[1] == "PO Number *"
    assert headers[2] == "Contract Name *"
    assert headers[7] == "Currency"


def test_generate_template_line_items_has_20_columns():
    from app.api.v1.routes.onboarding import _generate_bulk_template
    from openpyxl import load_workbook
    db_lists = {
        "license_types": [], "metrics": [], "regions": [],
        "categories": [], "sub_categories": [], "discovery_sources": [], "usage_methods": [],
    }
    data = _generate_bulk_template(db_lists)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Contract Line Items"]
    # Row 2 has column headers
    headers = [ws.cell(row=2, column=c).value for c in range(1, 21)]
    assert headers[0] == "Software Name *"
    assert headers[19] == "Usage Update Method"
